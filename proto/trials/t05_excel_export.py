"""T05 — Excel export: pivoted budget with a window, and the full ledger. No LLM."""
import io
import json
import sys
import urllib.request

sys.path.insert(0, __file__.rsplit("/", 1)[0])
from openpyxl import load_workbook
from runner import BASE, FAILS, expect, finish, reset

reset()
ops = [
    {"op": "create_book", "start": "2026-01-01", "end": "2027-01-01",
     "opening_balance": "1000", "grain": "month"},
    {"op": "add_item", "item": {"id": "salary", "name": "Salary", "kind": "flow",
     "direction": "in", "tags": {"cf": "cash"},
     "segments": [{"start": "2026-01-01", "recurrence": {"every": 1, "unit": "month"},
                   "amount": {"constant": "2000"}}], "settlement": "immediate"}},
    {"op": "add_item", "item": {"id": "rent", "name": "Rent", "kind": "flow",
     "direction": "out", "tags": {"cf": "cash"},
     "segments": [{"start": "2026-01-01", "recurrence": {"every": 1, "unit": "month"},
                   "amount": {"constant": "-800"}}], "settlement": "immediate"}},
    {"op": "add_event", "event": {"id": "ev-1", "date": "2026-02-10",
     "amount": "-300.00", "status": "forecast", "item": "rent"}},
]
req = urllib.request.Request(BASE + "/api/ops", data=json.dumps({"ops": ops}).encode(),
                             headers={"Content-Type": "application/json"})
out = json.load(urllib.request.urlopen(req))
bad = [d for r in out["reports"] for d in r["diagnostics"] if d["severity"] == "error"]
expect("setup ops clean", len(bad), 0)

# -- budget export: 6 months from March ------------------------------------- #
data = urllib.request.urlopen(BASE + "/api/export?mode=budget&months=6&start=2026-03").read()
wb = load_workbook(io.BytesIO(data))
ws = wb["Budget"]
rows = list(ws.values)
header = rows[0]
expect("window is 6 months", len(header) - 2, 6)
expect("window starts Mar", header[2] == "2026-03", True, tol=0)
by_name = {r[0]: r for r in rows if r and r[0]}
expect("salary Mar", by_name["Salary"][2], 2000)
expect("rent Mar", by_name["Rent"][2], -800)
closing_row = by_name.get("Closing balance")
# Jan: 1000+1200=2200, Feb: +1200-300=3100, Mar: 4300
expect("closing Mar in sheet", closing_row[2], 4300)

# -- ledger export ----------------------------------------------------------- #
data = urllib.request.urlopen(BASE + "/api/export?mode=ledger").read()
wb = load_workbook(io.BytesIO(data))
ws = wb["Ledger"]
rows = list(ws.values)
expect("ledger has header+1 row", len(rows), 2)
print("  ledger columns:", rows[0])
print("  ledger row:", rows[1])
finish("T05")
