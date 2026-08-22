"""T07 — initialize from a realistic human budget sheet.

Human conventions on purpose: month-name headers, expenses entered POSITIVE,
section headers, SUM/net/balance formula rows, a starting-balance corner cell,
a 13th-month salary, bimonthly utilities, one annual premium, a price rise.
Expected closing Dec = 13390 (hand-computed in comments below).
"""
import io
import sys

sys.path.insert(0, __file__.rsplit("/", 1)[0])
from openpyxl import Workbook
from runner import cell, closing, expect, find, finish, reset, state, upload

MODEL = sys.argv[1] if len(sys.argv) > 1 else "lite"

wb = Workbook()
ws = wb.active
ws.title = "Budget 2026"
ws["A1"] = "Family budget 2026"
ws["A3"] = "Starting balance:"
ws["B3"] = 2500

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
ws.append([])
ws.append([""] + months)                                   # row 5
ws.append(["INCOME"])                                      # row 6
ws.append(["Marco salary (net)"] + [2100] * 11 + [4200])   # 13th month in Dec
ws.append(["Anna part-time"] + [950] * 8 + [0, 400, 400, 400])
ws.append(["Child benefit"] + [260] * 12)
ws.append(["Total income"] + [f"=SUM({c}7:{c}9)" for c in "BCDEFGHIJKLM"])
ws.append(["EXPENSES"])
ws.append(["Mortgage"] + [1150] * 12)
ws.append(["Groceries"] + [780] * 12)
ws.append(["Utilities (bimonthly)"] + [190 if i % 2 == 0 else 0 for i in range(12)])
ws.append(["Car insurance (annual)"] + [0] * 5 + [540] + [0] * 6)
ws.append(["Kindergarten"] + [310] * 7 + [0] + [330] * 4)
ws.append(["Total expenses"] + [f"=SUM({c}13:{c}17)" for c in "BCDEFGHIJKLM"])
ws.append(["Net"] + [f"={c}10-{c}18" for c in "BCDEFGHIJKLM"])
ws.append(["End balance", "=$B$3+B19"] + [f"={c}20+{chr(ord(c)+1)}19" for c in "BCDEFGHIJKL"])
buf = io.BytesIO()
wb.save(buf)

reset()
upload(buf.getvalue(), "family-budget-2026.xlsx", model=MODEL)
s = state()
if not s.get("book"):
    print("  no book created — FAIL")
    finish("T07 " + MODEL)
    sys.exit(1)
print("  items:", {i["id"]: i["name"] for i in s["items"]})
print("  opening:", s["book"]["opening_balance"])

salary = find(s, "marco") if find(s, "marco") != "?" else find(s, "salary")
anna = find(s, "anna") if find(s, "anna") != "?" else find(s, "part")
util = find(s, "utilit")
car = find(s, "car") if find(s, "car") != "?" else find(s, "insurance")
kind = find(s, "kinder")
expect("opening 2500", float(s["book"]["opening_balance"]), 2500)
expect("salary Nov", cell(s, salary, "2026-11"), 2100)
expect("salary Dec (13th)", cell(s, salary, "2026-12"), 4200)
expect("anna Sep = 0", cell(s, anna, "2026-09") or 0, 0)
expect("anna Nov", cell(s, anna, "2026-11"), 400)
expect("utilities Feb = 0", cell(s, util, "2026-02") or 0, 0)
expect("utilities Mar", cell(s, util, "2026-03"), -190)
expect("car insurance Jun", cell(s, car, "2026-06"), -540)
expect("car insurance Jul = 0", cell(s, car, "2026-07") or 0, 0)
expect("kindergarten Aug = 0", cell(s, kind, "2026-08") or 0, 0)
expect("kindergarten Sep", cell(s, kind, "2026-09"), -330)
# Jan: 2100+950+260 - (1150+780+190+310) = 880 -> 3380
expect("closing Jan", closing(s, "2026-01"), 3380)
# year: income 27300+8800+3120=39220; expenses 13800+9360+1140+540+3490=28330 -> +10890
expect("closing Dec", closing(s, "2026-12"), 13390)
finish("T07 " + MODEL)
