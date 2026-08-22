"""CashKit proto webapp: natural language -> CashKit SDK operations.

Run:  uv run uvicorn proto.server:app --port 8765 --reload
ponytail: one process, one book, module-global kit. Multi-book/multi-user is app-layer
work the prototype does not need.
"""
from __future__ import annotations

import io
import json
import shutil
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from cashkit.model import Event, Grain, Item, PeriodRange
from cashkit.sdk import CashKit, balance_series, create_book

from . import llm

ROOT = Path(__file__).parent
BOOK_DIR = ROOT / "books" / "current"
SCALE = Decimal(10000)  # engine columns are int64 minor units @ 4dp

app = FastAPI()


class S:
    kit: CashKit | None = None


def _open() -> CashKit | None:
    if S.kit is None and BOOK_DIR.exists():
        kit, _diags = CashKit.open(BOOK_DIR)
        S.kit = kit
    return S.kit


def _dec(v: Any) -> Decimal:
    return Decimal(str(v))


def _money(units: int) -> str:
    return str(Decimal(int(units)) / SCALE)


def _diag(d) -> dict:
    return {"code": d.code, "severity": d.severity, "message": d.message,
            "fix": d.suggested_fix, "item": d.item_id}


def _err(msg: str) -> dict:
    return {"code": "PROTO-EXC", "severity": "error", "message": msg, "fix": "", "item": None}


def _norm_settlement(s: Any) -> Any:
    if s is None or s == "immediate":
        return {"due": [{"share": "1", "offset": "0d"}]}
    if isinstance(s, dict) and "net" in s:
        n = str(s["net"]).strip()
        return {"due": [{"share": "1", "offset": n if n[-1:] in "dwmy" else n + "d"}]}
    if isinstance(s, dict):
        for term in s.get("due") or []:
            off = term.get("offset") if isinstance(term, dict) else None
            if isinstance(off, int) or (isinstance(off, str) and off.isdigit()):
                term["offset"] = f"{off}d"  # bare number means days
    return s


def apply_op(op: dict) -> dict:
    """Apply one op; every failure comes back as diagnostics, never an exception."""
    diags: list[dict] = []
    try:
        kind = op.get("op")
        if kind == "create_book":
            S.kit = None
            shutil.rmtree(BOOK_DIR, ignore_errors=True)
            ref = create_book(
                BOOK_DIR,
                id=op.get("id", "proto"),
                horizon=PeriodRange(start=date.fromisoformat(op["start"]),
                                    end=date.fromisoformat(op["end"])),
                opening_balance=_dec(op["opening_balance"]),
                grain=Grain(op.get("grain", "month")),
                params={k: _dec(v) for k, v in (op.get("params") or {}).items()},
            )
            S.kit = ref.kit
            diags += [_diag(d) for d in ref.diagnostics]
        elif kind == "add_item":
            d = dict(op["item"])
            if d.get("kind", "flow") == "flow":
                d["settlement"] = _norm_settlement(d.get("settlement"))
            for seg in d.get("segments") or []:
                # a schedule IS the occurrence series; the required recurrence is
                # semantically inert there, so default it instead of bouncing the op
                if (isinstance(seg, dict) and "recurrence" not in seg
                        and "schedule" in (seg.get("amount") or {})):
                    seg["recurrence"] = {"every": 1, "unit": "month"}
            item = Item.model_validate(d)
            r = _require().add_item(item)
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "add_derived":
            kw = {}
            if op.get("kind") == "stock":
                kw["kind"] = "stock"
            if op.get("name"):
                kw["name"] = op["name"]
            if op.get("direction"):
                kw["direction"] = op["direction"]
            r = _require().add_derived(op["id"], op["formula"], op.get("tags"), **kw)
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "set_param":
            r = _require().set_param(op["key"], _dec(op["value"]))
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "add_event":
            e = Event.model_validate(op["event"])
            r = _require().add_event(e)
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "fork_scenario":
            r = _require().scenarios.fork(op.get("parent", "base"), op["id"], note=op.get("note", ""))
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "set_scenario_param":
            r = _require().scenarios.set_param(op["scenario"], op["key"], _dec(op["value"]))
            diags += [_diag(x) for x in r.diagnostics]
        elif kind == "retag":
            r = _require().retag(op["selector"], op["tags"])
            diags += [_diag(x) for x in getattr(r, "diagnostics", ())]
        else:
            diags.append(_err(f"unknown op {kind!r}"))
    except ValidationError as exc:
        parts = [f"{'.'.join(map(str, e['loc']))}: {e['msg']} (input {e.get('input')!r})"
                 for e in exc.errors(include_url=False)]
        diags.append(_err("invalid payload: " + "; ".join(parts)[:1500]))
    except Exception as exc:  # noqa: BLE001 — surface everything to the LLM loop
        diags.append(_err(f"{type(exc).__name__}: {exc}"[:800]))
    return {"op": op, "ok": not any(d["severity"] == "error" for d in diags),
            "diagnostics": diags}


def _require() -> CashKit:
    kit = _open()
    if kit is None:
        raise RuntimeError("no book exists yet — emit create_book first")
    return kit


# -- state ----------------------------------------------------------------- #

def book_state_json() -> str:
    """Compact state handed to the LLM."""
    kit = _open()
    if kit is None:
        return '{"book": null}'
    book = kit.scenarios.resolve("base")
    items = []
    for it in book.items.values():
        e: dict[str, Any] = {"id": it.id, "kind": it.kind, "name": it.name}
        if it.direction:
            e["direction"] = it.direction
        if it.tags:
            e["tags"] = dict(it.tags)
        if it.formula:
            e["formula"] = it.formula
        segs = []
        for s in it.segments:
            seg = {"start": str(s.start), "end": str(s.end) if s.end else None,
                   "every": f"{s.recurrence.every}{s.recurrence.unit.value[0]}"}
            if s.amount.constant is not None:
                seg["constant"] = str(s.amount.constant)
            else:
                seg["schedule"] = [[str(d), str(a)] for d, a in (s.amount.schedule or [])]
            segs.append(seg)
        if segs:
            e["segments"] = segs
        items.append(e)
    state = {
        "book": {
            "horizon": [str(book.horizon.start), str(book.horizon.end)],
            "grain": book.base_grain.value,
            "opening_balance": str(book.opening_balance),
            "params": {k: str(v) for k, v in book.params.items()},
            "scenarios": [s for s in _scenario_ids(kit)],
        },
        "items": items,
    }
    return json.dumps(state, separators=(",", ":"))


def _scenario_ids(kit: CashKit) -> list[str]:
    try:
        return list(kit.scenarios.ids())  # type: ignore[attr-defined]
    except Exception:
        return ["base"]


def run_payload(scenario: str = "base") -> dict:
    kit = _open()
    if kit is None:
        return {"book": None}
    run = kit.run(scenario)
    starts = run.result.periods.starts
    months = [d.isoformat()[:7] for d in starts]
    rows = []
    for it in run.book.items.values():
        try:
            cash = [_money(v) for v in run.result.cash[it.id]]
            accrual = [_money(v) for v in run.result.accrual[it.id]]
        except KeyError:
            continue
        rows.append({"id": it.id, "name": it.name, "kind": it.kind,
                     "direction": it.direction, "cash": cash, "value": accrual})
    series, source = balance_series(run.result, run.book)
    summary = run.summary()
    return {
        "book": {"id": run.book.id, "opening_balance": str(run.book.opening_balance),
                 "params": {k: str(v) for k, v in run.book.params.items()}},
        "scenario": scenario,
        "months": months,
        "items": rows,
        "closing": [_money(v) for v in series],
        "summary": {
            "min_cash": str(summary.min_cash), "min_cash_period": str(summary.min_cash_period),
            "closing_balance": str(summary.closing_balance),
            "total_inflow": str(summary.total_inflow),
            "total_outflow": str(summary.total_outflow),
            "net_cash": str(summary.net_cash),
            "runway_end": str(summary.runway_end) if summary.runway_end else None,
        },
        "diagnostics": [_diag(d) for d in run.diagnostics],
    }


# -- LLM loop ---------------------------------------------------------------- #

def llm_ops_loop(messages: list[dict], model_key: str, max_rounds: int = 3) -> dict:
    model = llm.MODELS.get(model_key, model_key)
    rounds = []
    reply = ""
    temperature = 0.0
    for _ in range(max_rounds):
        raw = llm.complete(messages, model=model, temperature=temperature)
        try:
            parsed = llm.extract_json(raw)
        except (ValueError, json.JSONDecodeError) as exc:
            temperature = 0.7  # a temp-0 retry reproduces the same broken JSON
            rounds.append({"raw": raw[:2000], "ops": [], "reports": [],
                           "error": f"unparseable output: {exc}"})
            messages += [{"role": "assistant", "content": raw},
                         {"role": "user", "content":
                          f"Your output was not a valid JSON object ({exc}). "
                          "Return ONLY the JSON object, no fences, no prose."}]
            continue
        ops = parsed.get("ops", []) or []
        reply = parsed.get("reply", "")
        reports = [apply_op(o) for o in ops]
        if S.kit is not None:
            S.kit.save()
        errors = [d for r in reports for d in r["diagnostics"] if d["severity"] == "error"]
        run_diags: list[dict] = []
        if S.kit is not None and not errors:
            run_diags = [d for d in run_payload()["diagnostics"] if d["severity"] == "error"]
        rounds.append({"ops": ops, "reports": reports, "reply": reply,
                       "run_errors": run_diags})
        problems = errors + run_diags
        if not problems:
            break
        messages += [{"role": "assistant", "content": raw},
                     {"role": "user", "content":
                      "Some operations failed. Diagnostics:\n" + json.dumps(problems)
                      + "\nCurrent book state:\n" + book_state_json()
                      + "\nReturn a JSON object with ONLY the ops that fix these "
                        "problems (already-applied ops stay applied; re-emitting an "
                        "add_item replaces that item)."}]
    return {"reply": reply, "rounds": rounds}


# -- routes ------------------------------------------------------------------ #

@app.get("/")
async def index():
    return FileResponse(ROOT / "index.html")


@app.get("/api/state")
async def state(scenario: str = "base"):
    return run_payload(scenario)


@app.post("/api/reset")
async def reset():
    S.kit = None
    shutil.rmtree(BOOK_DIR, ignore_errors=True)
    return {"ok": True}


@app.post("/api/ops")
async def raw_ops(body: dict):
    """Dev/trial hook: apply ops directly, no LLM."""
    reports = [apply_op(o) for o in body.get("ops", [])]
    if S.kit is not None:
        S.kit.save()
    return {"reports": reports, "state": run_payload()}


@app.post("/api/chat")
async def chat(body: dict):
    history = body.get("history", [])[-10:]
    messages = llm.chat_system(book_state_json()) + history + [
        {"role": "user", "content": body["message"]}]
    out = llm_ops_loop(messages, body.get("model", llm.DEFAULT))
    out["state"] = run_payload()
    return JSONResponse(out)


@app.post("/api/upload")
async def upload(file: UploadFile = File(...), model: str = Form(llm.DEFAULT)):
    data = await file.read()
    grid = sheet_text(data)
    messages = llm.upload_system(book_state_json()) + [
        {"role": "user", "content": "Spreadsheet contents:\n" + grid}]
    out = llm_ops_loop(messages, model)
    out["state"] = run_payload()
    return JSONResponse(out)


@app.get("/api/export")
async def export(mode: str = "budget", months: int = 12, start: str | None = None,
           scenario: str = "base"):
    kit = _open()
    if kit is None:
        return JSONResponse({"error": "no book"}, status_code=400)
    wb = Workbook()
    ws = wb.active
    if mode == "ledger":
        ws.title = "Ledger"
        table = kit.query_events()
        ws.append(list(table.columns))

        def xl(v):
            if v is None or isinstance(v, (int, float, str, date)):
                return v
            if isinstance(v, Decimal):
                return float(v)
            return json.dumps(v, default=str)

        for row in table.rows:
            ws.append([xl(v) for v in row])
    else:
        payload = run_payload(scenario)
        months = max(1, min(int(months), len(payload["months"])))
        lo = 0
        if start:
            lo = next((i for i, m in enumerate(payload["months"]) if m >= start[:7]), 0)
        hi = min(lo + months, len(payload["months"]))
        ws.title = "Budget"
        ws.append(["Item", "Kind"] + payload["months"][lo:hi])
        ws.append(["Opening balance", "meta", float(payload["book"]["opening_balance"])])
        for it in payload["items"]:
            vals = it["cash"] if it["kind"] != "stock" else it["value"]
            ws.append([it["name"], it["kind"]] + [float(v) for v in vals[lo:hi]])
        ws.append([])
        ws.append(["Closing balance", ""] + [float(v) for v in payload["closing"][lo:hi]])
    buf = io.BytesIO()
    wb.save(buf)
    name = f"cashkit-{mode}.xlsx"
    return Response(buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={name}"})


@app.get("/api/usage")
async def usage():
    spent = 0.0
    calls = 0
    if llm.USAGE_LOG.exists():
        for line in llm.USAGE_LOG.read_text().splitlines():
            rec = json.loads(line)
            spent += rec.get("cost") or 0.0
            calls += 1
    try:
        ks = llm.key_status()
        remaining = ks.get("limit_remaining")
    except Exception:
        remaining = None
    return {"calls": calls, "cost": round(spent, 4), "limit_remaining": remaining}


def sheet_text(data: bytes, max_chars: int = 15000) -> str:
    """Values + formulas of every sheet, compact, capped."""
    vals = load_workbook(io.BytesIO(data), data_only=True)
    fx = load_workbook(io.BytesIO(data), data_only=False)
    out: list[str] = []
    for name in vals.sheetnames:
        out.append(f"## sheet {name}")
        wsv, wsf = vals[name], fx[name]
        for rv, rf in zip(wsv.iter_rows(max_row=300, max_col=40),
                          wsf.iter_rows(max_row=300, max_col=40)):
            cells = []
            for cv, cf in zip(rv, rf):
                if cv.value is None and cf.value is None:
                    continue
                s = f"{cv.coordinate}={cv.value!r}"
                if isinstance(cf.value, str) and cf.value.startswith("="):
                    s += f" [{cf.value}]"
                cells.append(s)
            if cells:
                out.append(" ".join(cells))
    text = "\n".join(out)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...[truncated]"
    return text
