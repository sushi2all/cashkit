"""Reading a workbook: values, formulas, and the sheet's own total rows.

Two loads, as ``proto/server.py:sheet_text`` does: one ``data_only`` load for
the cached values Excel computed, and one formula load, so the model sees both
``B14=1234.5`` and ``[=SUM(B3:B13)]``. A formula is what tells a reader that a
row is a subtotal rather than another line item, and that is exactly what the
reconciliation loop needs.

**Where the float boundary is.** A spreadsheet cell is a float — that is
Excel's type, not a choice this service makes. The conversion happens exactly
once, here, at the cell boundary, through ``str()`` so the shortest exact
repr is what becomes the ``Decimal`` (the mirror of the export's D-MLP-13
conversion). No arithmetic is ever done on a float in this module.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

#: A workbook bigger than this in either direction is read up to the bound. A
#: personal budget is far inside it, and an unbounded read is unbounded work on
#: the event-loop thread.
MAX_ROWS = 400
MAX_COLS = 60

#: How much sheet text one prompt carries. The proto's number.
MAX_PROMPT_CHARS = 15_000

#: Row labels that mark a sheet's own arithmetic — the rows a reconciliation
#: checks against. Matched case-insensitively, as whole words, and in the four
#: languages a European household budget is likely to be written in.
TOTAL_WORDS = (
    "total", "totale", "totaal", "totaux", "somme", "summe", "gesamt",
    "subtotal", "subtotale", "sub-total",
    "net", "netto", "netta", "saldo", "balance", "closing", "cumulative",
    "sum", "difference", "surplus", "savings", "risparmio",
)

_WORD = re.compile(r"[a-zà-ÿ]+")


class UnreadableWorkbook(ValueError):
    """The upload was not a workbook this service can read."""


@dataclass(frozen=True)
class Cell:
    """One cell, in both of its readings."""

    sheet: str
    coordinate: str
    value: Any
    formula: str | None

    @property
    def ref(self) -> str:
        return f"{self.sheet}!{self.coordinate}"


@dataclass
class Sheets:
    """A parsed workbook, addressable by ``Sheet!A1``."""

    names: list[str]
    cells: dict[str, dict[str, Cell]] = field(default_factory=dict)

    def cell(self, ref: str) -> Cell | None:
        sheet, coordinate = split_ref(ref, default=self.names[0] if self.names else "")
        return self.cells.get(sheet, {}).get(coordinate.upper())

    def row(self, sheet: str, row: int) -> list[Cell]:
        return sorted(
            (c for c in self.cells.get(sheet, {}).values() if row_of(c.coordinate) == row),
            key=lambda c: column_index(c.coordinate),
        )

    @property
    def filled_cells(self) -> int:
        return sum(len(s) for s in self.cells.values())


def split_ref(ref: str, *, default: str = "") -> tuple[str, str]:
    """``"Budget!B14"`` → ``("Budget", "B14")``. A bare ``"B14"`` uses ``default``."""
    text = str(ref).strip().replace("$", "")
    if "!" in text:
        sheet, _, coordinate = text.rpartition("!")
        return sheet.strip().strip("'"), coordinate.strip()
    return default, text


def column_index(coordinate: str) -> int:
    letters = "".join(ch for ch in coordinate if ch.isalpha()).upper()
    index = 0
    for ch in letters:
        index = index * 26 + (ord(ch) - 64)
    return index


def row_of(coordinate: str) -> int:
    digits = "".join(ch for ch in coordinate if ch.isdigit())
    return int(digits) if digits else 0


def parse(data: bytes) -> Sheets:
    """Read every sheet's values and formulas."""
    try:
        values = load_workbook(io.BytesIO(data), data_only=True)
        formulas = load_workbook(io.BytesIO(data), data_only=False)
    except Exception as exc:  # noqa: BLE001 — any openpyxl failure is one thing to the user
        raise UnreadableWorkbook(f"{type(exc).__name__}: {exc}") from exc

    sheets = Sheets(names=list(values.sheetnames))
    for name in sheets.names:
        table: dict[str, Cell] = {}
        wsv, wsf = values[name], formulas[name]
        rows_f = wsf.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS)
        for row_v, row_f in zip(wsv.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS), rows_f):
            for cell_v, cell_f in zip(row_v, row_f):
                formula = cell_f.value if isinstance(cell_f.value, str) and cell_f.value.startswith("=") else None
                if cell_v.value is None and formula is None:
                    continue
                table[cell_v.coordinate.upper()] = Cell(
                    sheet=name,
                    coordinate=cell_v.coordinate.upper(),
                    value=cell_v.value,
                    formula=formula,
                )
        sheets.cells[name] = table
    return sheets


def as_decimal(value: Any) -> Decimal | None:
    """The one float→Decimal boundary. Anything not a number returns ``None``."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):
        # str() of a float is its shortest exact repr, which is what the person
        # who typed the cell meant. No arithmetic happens on the float itself.
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip().replace(" ", " ")
        text = re.sub(r"[€$£\s]", "", text)
        if not text:
            return None
        negative = text.startswith("(") and text.endswith(")")
        if negative:
            text = text[1:-1]
        # A European sheet writes 1.234,56 and an English one 1,234.56.
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".") if len(text.split(",")[-1]) != 3 else text.replace(",", "")
        try:
            number = Decimal(text)
        except (InvalidOperation, ValueError):
            return None
        return -number if negative else number
    return None


def read_number(sheets: Sheets, ref: str) -> Decimal | None:
    """The value at ``Sheet!B14``, as a Decimal, or ``None`` if there is none."""
    cell = sheets.cell(ref)
    return None if cell is None else as_decimal(cell.value)


def sheet_text(sheets: Sheets, *, max_chars: int = MAX_PROMPT_CHARS) -> str:
    """Values and formulas of every sheet, compact and capped.

    The shape is the proto's: one line per spreadsheet row, each filled cell as
    ``COORD=value`` with ``[=FORMULA]`` appended where there is one.
    """
    out: list[str] = []
    for name in sheets.names:
        out.append(f"## sheet {name}")
        table = sheets.cells.get(name, {})
        by_row: dict[int, list[Cell]] = {}
        for cell in table.values():
            by_row.setdefault(row_of(cell.coordinate), []).append(cell)
        for row in sorted(by_row):
            cells = sorted(by_row[row], key=lambda c: column_index(c.coordinate))
            rendered = " ".join(
                f"{c.coordinate}={c.value!r}" + (f" [{c.formula}]" if c.formula else "")
                for c in cells
            )
            out.append(rendered)
    text = "\n".join(out)
    if len(text) > max_chars:
        text = text[:max_chars] + "\n...[truncated]"
    return text


def total_row_candidates(sheets: Sheets) -> list[dict[str, Any]]:
    """The sheet's own subtotal, total and balance cells (SPEC §7.2).

    Deterministic, host-side, and offered to the model as *candidates* — the
    model says what each one means, and the host reads its value. A row counts
    when its label contains one of :data:`TOTAL_WORDS` as a whole word, or when
    a numeric cell on it carries a ``SUM`` formula.
    """
    found: list[dict[str, Any]] = []
    for name in sheets.names:
        table = sheets.cells.get(name, {})
        by_row: dict[int, list[Cell]] = {}
        for cell in table.values():
            by_row.setdefault(row_of(cell.coordinate), []).append(cell)
        for row in sorted(by_row):
            cells = sorted(by_row[row], key=lambda c: column_index(c.coordinate))
            label = next((str(c.value) for c in cells if isinstance(c.value, str)), "")
            words = set(_WORD.findall(label.lower()))
            labelled = bool(words & set(TOTAL_WORDS))
            for cell in cells:
                if as_decimal(cell.value) is None:
                    continue
                summed = bool(cell.formula and "SUM" in cell.formula.upper())
                if not (labelled or summed):
                    continue
                found.append(
                    {
                        "ref": cell.ref,
                        "label": label.strip() or cell.ref,
                        "value": str(as_decimal(cell.value)),
                        "formula": cell.formula,
                    }
                )
    return found


def header_hints(sheets: Sheets) -> list[str]:
    """The first text row of each sheet, which is usually the month header."""
    hints: list[str] = []
    for name in sheets.names:
        table = sheets.cells.get(name, {})
        rows = sorted({row_of(c.coordinate) for c in table.values()})
        for row in rows[:4]:
            cells = sheets.row(name, row)
            if sum(1 for c in cells if isinstance(c.value, str)) >= 2:
                hints.append(f"{name} row {row}: " + " | ".join(str(c.value) for c in cells))
                break
    return hints


__all__ = [
    "Cell",
    "MAX_PROMPT_CHARS",
    "Sheets",
    "UnreadableWorkbook",
    "as_decimal",
    "column_index",
    "header_hints",
    "parse",
    "read_number",
    "row_of",
    "sheet_text",
    "split_ref",
    "total_row_candidates",
]
