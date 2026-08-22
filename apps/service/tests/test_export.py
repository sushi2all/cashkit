"""``GET /export`` — the workbook (SPEC §3)."""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import load_workbook

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def test_budget_export_is_a_workbook(seeded_client):
    response = await seeded_client.get("/export", params={"mode": "budget", "months": 12})
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX

    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet.title == "Budget"
    header = [c.value for c in sheet[1]]
    assert header[:2] == ["Item", "Kind"]
    assert header[2] == "2026-01"
    assert len(header) == 14

    names = [row[0].value for row in sheet.iter_rows()]
    assert "Opening balance" in names and "Closing balance" in names
    assert {"Rent", "Salary", "Insurance"} <= set(names)


async def test_exported_figures_match_the_api(seeded_client):
    api = (await seeded_client.get("/book/state")).json()
    response = await seeded_client.get("/export", params={"mode": "budget", "months": 12})
    sheet = load_workbook(io.BytesIO(response.content)).active

    closing = next(row for row in sheet.iter_rows() if row[0].value == "Closing balance")
    exported = [c.value for c in closing[2:]]
    expected = [float(Decimal(m["exact"])) for m in api["closing"]]
    assert exported == expected


async def test_ledger_export_carries_the_events(seeded_client):
    response = await seeded_client.get("/export", params={"mode": "ledger"})
    sheet = load_workbook(io.BytesIO(response.content)).active
    assert sheet.title == "Ledger"
    header = [c.value for c in sheet[1]]
    assert header[:4] == ["id", "date", "amount", "status"]
    statuses = {row[3].value for row in sheet.iter_rows(min_row=2)}
    assert "actual" in statuses


async def test_the_window_can_be_narrowed(seeded_client):
    response = await seeded_client.get(
        "/export", params={"mode": "budget", "months": 3, "start": "2026-04-01"}
    )
    sheet = load_workbook(io.BytesIO(response.content)).active
    header = [c.value for c in sheet[1]]
    assert header[2:] == ["2026-04", "2026-05", "2026-06"]


async def test_export_needs_a_session(client):
    assert (await client.get("/export")).status_code == 401
