"""T05 — exporting the book the model built (proto T05).

    Tried: build a book, export the budget window, parse it and check cells.

Proto T05 needed no model at all; its service equivalent still needs one,
because the book under export is the one a turn authored. The export itself is
deterministic, which is the point: the same figures the API serves are the
figures in the sheet.
"""

from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from trials.live import author, make_book, state

pytestmark = pytest.mark.live_model

OPENING = Decimal("150.00")
BUILD = (
    "My scholarship pays me 700 a month all year from January, and my rent is "
    "455 a month all year from January."
)


async def test_the_export_matches_what_the_api_serves(live_session):
    await make_book(live_session, str(OPENING))
    await author(live_session, BUILD)

    response = await live_session.get("/export", params={"mode": "budget", "months": 12})
    assert response.status_code == 200
    sheet = load_workbook(io.BytesIO(response.content)).active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][0] == "Item" and rows[0][2] == "2026-01"
    assert rows[1][0] == "Opening balance"
    assert Decimal(str(rows[1][2])) == OPENING

    closing_row = next(r for r in rows if r[0] == "Closing balance")
    body = await state(live_session)
    for index, cell in enumerate(closing_row[2:14]):
        assert Decimal(str(cell)) == Decimal(body["closing"][index]["exact"])


async def test_the_ledger_export_works_on_a_book_with_no_events(live_session):
    await make_book(live_session, str(OPENING))
    await author(live_session, BUILD)

    response = await live_session.get("/export", params={"mode": "ledger"})
    assert response.status_code == 200
    assert load_workbook(io.BytesIO(response.content)).active.title == "Ledger"
