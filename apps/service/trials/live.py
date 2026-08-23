"""Fixtures and helpers for the trials that call the real model.

Every trial in ``t01_…`` to ``t12_…`` runs against the pinned model
(``google/gemini-3.7-flash``, ADR-0028) through the real service, and asserts
the **final state of the book numerically**. That is what makes them the
model-behaviour gate (SPEC §10): a prompt change or a model change reruns them,
and a plausible-looking answer with the wrong number fails.

Three rules hold in every trial here:

* the clock is frozen (D-MLP-12), so no assertion depends on the day it runs;
* a proposal is accepted **through the real endpoint**, never around it — the
  trial is a user with a keyboard, not a shortcut into the applier;
* the expected figures are computed in the trial from the sentence the user
  said, in :class:`~decimal.Decimal`, so the assertion is an independent check
  and not a copy of what the engine happened to produce.

They are marked ``live_model`` and excluded from a per-commit run (SPEC §10).
Run them with ``uv run pytest apps/service/trials -m live_model``.
"""

from __future__ import annotations

import os
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient

from cashkit_service.agent.transport import OpenRouterTransport
from cashkit_service.app import create_app
from cashkit_service.books import BookRuntime

REPO_ROOT = Path(__file__).resolve().parents[3]

#: The whole horizon every trial book uses, so "December" means one thing.
HORIZON = {"horizon_start": "2026-01-01", "horizon_end": "2027-01-01"}


def api_key() -> str | None:
    """The OpenRouter key, from the environment or the repo-root ``.env``."""
    for name in ("CASHKIT_LLM_API_KEY", "OPENROUTER_API_KEY"):
        if os.environ.get(name):
            return os.environ[name]
    env = REPO_ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            if line.startswith("OPENROUTER_API_KEY="):
                return line.split("=", 1)[1].strip()
    return None


@pytest.fixture
def live_transport(settings):
    key = api_key()
    if not key:
        pytest.skip("no OPENROUTER_API_KEY: the live trials need the pinned model")
    return OpenRouterTransport(
        api_key=key,
        model=settings.llm_model,
        base_url=settings.llm_base_url,
        timeout_seconds=settings.llm_timeout_seconds,
        max_tokens=settings.llm_max_tokens,
    )


@pytest.fixture
def live_app(settings, clock, mailer, database, books_root, live_transport):
    return create_app(
        settings=settings,
        clock=clock,
        mailer=mailer,
        database=database,
        book_runtime=BookRuntime(books_root),
        transport=live_transport,
    )


@pytest_asyncio.fixture
async def live_session(live_app, mailer) -> AsyncClient:
    """An authenticated client on the live service, with no book yet."""
    async with AsyncClient(
        transport=ASGITransport(app=live_app), base_url="http://test"
    ) as client:
        email = "trial@example.com"
        await client.post("/auth/link", json={"email": email, "platform": "web"})
        token = (
            await client.post(
                "/auth/verify", json={"token": mailer.last_for(email).token, "platform": "web"}
            )
        ).json()["token"]
        client.headers["Authorization"] = f"Bearer {token}"
        yield client
    live_app.state.books.close_all()


async def new_session(live_app, mailer, email: str) -> AsyncClient:
    """A second account on the same service — one book per user (ADR-0027)."""
    client = AsyncClient(transport=ASGITransport(app=live_app), base_url="http://test")
    await client.post("/auth/link", json={"email": email, "platform": "web"})
    token = (
        await client.post(
            "/auth/verify", json={"token": mailer.last_for(email).token, "platform": "web"}
        )
    ).json()["token"]
    client.headers["Authorization"] = f"Bearer {token}"
    return client


async def sheet_text(client: AsyncClient, **params: Any) -> str:
    """The exported budget workbook, rendered as the text a user would paste.

    The xlsx parsing pipeline is S5's (SPEC §7). What belongs to S2 is the
    model-behaviour half of proto T06/T07: can the model turn a tabular budget
    back into a correct book? Rendering our own export as text asks exactly that
    question without building the import loop (D-MLP-27).
    """
    import io

    from openpyxl import load_workbook

    response = await client.get("/export", params={"mode": "budget", "months": 12, **params})
    assert response.status_code == 200, response.text
    sheet = load_workbook(io.BytesIO(response.content)).active
    lines = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else (f"{c:g}" if isinstance(c, float) else str(c)) for c in row]
        lines.append(" | ".join(cells).rstrip(" |"))
    return "\n".join(line for line in lines if line.strip())


async def make_book(client: AsyncClient, opening_balance: str, **overrides: Any) -> None:
    response = await client.post(
        "/books", json={**HORIZON, "opening_balance": opening_balance, **overrides}
    )
    assert response.status_code == 201, response.text


# --- driving a turn -------------------------------------------------------- #


async def turn(client: AsyncClient, text: str, **body: Any) -> dict[str, Any]:
    response = await client.post("/turns", json={"text": text, **body})
    assert response.status_code == 200, response.text
    return response.json()


async def apply(client: AsyncClient, result: dict[str, Any]) -> dict[str, Any]:
    """Accept the turn's proposal through the real endpoint.

    Accepting programmatically is legitimate: it is a real accept through the
    real confirmation route, which is what a user's tap does. Constructing a
    change without a card is not, and T13 is the test that says so.
    """
    assert result["kind"] == "proposal", _explain(result)
    card = result["proposal"]
    assert not [d for d in card["diagnostics"] if d["severity"] == "error"], card["diagnostics"]
    response = await client.post(f"/proposals/{card['id']}", json={"action": "accept"})
    assert response.status_code == 200, response.text
    applied = response.json()
    assert applied["kind"] == "applied", applied
    return applied


async def author(client: AsyncClient, text: str, **body: Any) -> dict[str, Any]:
    """One authoring turn, applied. The commonest trial step."""
    result = await turn(client, text, **body)
    await apply(client, result)
    return result


def _explain(result: dict[str, Any]) -> str:
    return (
        f"kind={result['kind']} reply={result.get('reply')!r} "
        f"diagnostics={result.get('diagnostics')} "
        f"clarification={result.get('clarification')!r}"
    )


# --- reading the book back ------------------------------------------------- #


async def state(client: AsyncClient, scenario: str | None = None) -> dict[str, Any]:
    params = {"scenario": scenario} if scenario else None
    response = await client.get("/book/state", params=params)
    assert response.status_code == 200, response.text
    return response.json()


async def closing(client: AsyncClient, month: str, scenario: str | None = None) -> Decimal:
    """The closing balance at the end of ``month`` (``"2026-01"``)."""
    body = await state(client, scenario)
    index = [m[:7] for m in body["months"]].index(month)
    return Decimal(body["closing"][index]["exact"])


async def closings(client: AsyncClient, scenario: str | None = None) -> list[Decimal]:
    """The whole closing series, month by month."""
    body = await state(client, scenario)
    return [Decimal(c["exact"]) for c in body["closing"]]


async def events(client: AsyncClient) -> list[dict[str, Any]]:
    response = await client.get("/book/events")
    assert response.status_code == 200, response.text
    return response.json()["events"]


async def item_month(
    client: AsyncClient, item: str, month: str, measure: str = "cash",
    scenario: str | None = None,
) -> Decimal:
    body = await state(client, scenario)
    index = [m[:7] for m in body["months"]].index(month)
    series = next(i for i in body["items"] if i["id"] == item)
    return Decimal(series[measure][index]["exact"])


async def item_total(
    client: AsyncClient, item: str, measure: str = "cash", scenario: str | None = None
) -> Decimal:
    body = await state(client, scenario)
    series = next(i for i in body["items"] if i["id"] == item)
    return sum((Decimal(v["exact"]) for v in series[measure]), Decimal(0))


async def item_ids(client: AsyncClient, scenario: str | None = None) -> set[str]:
    return {i["id"] for i in (await state(client, scenario))["items"]}


async def total_by_direction(
    client: AsyncClient, direction: str, month: str | None = None,
    scenario: str | None = None,
) -> Decimal:
    """Every item's cash in one direction, summed. Used when ids are the model's."""
    body = await state(client, scenario)
    months = [m[:7] for m in body["months"]]
    total = Decimal(0)
    for series in body["items"]:
        if series["kind"] == "stock":
            continue
        for period, value in zip(months, series["cash"], strict=True):
            if month is not None and period != month:
                continue
            amount = Decimal(value["exact"])
            if (direction == "in" and amount > 0) or (direction == "out" and amount < 0):
                total += amount
    return total


def digits(text: str) -> str:
    """A reply's figures, with formatting removed, for a containment check."""
    return text.replace(",", "").replace(" ", "").replace(" ", "").replace(" ", "")


__all__ = [
    "HORIZON",
    "apply",
    "api_key",
    "author",
    "closing",
    "closings",
    "events",
    "digits",
    "item_ids",
    "item_month",
    "item_total",
    "live_app",
    "live_session",
    "live_transport",
    "make_book",
    "new_session",
    "sheet_text",
    "state",
    "total_by_direction",
    "turn",
]
